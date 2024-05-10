import pandas as pd
import tkinter as tk
import openpyxl


from pyexcel.cookbook import merge_all_to_a_book
# import pyexcel.ext.xlsx # no longer required if you use pyexcel >= 0.2.2 
import glob
merge_all_to_a_book(glob.glob("C:\Python\Repository\Mobax\Devices-05-10-2024.csv"), "C:\Python\Repository\Mobax\Device-Inventory.xlsx")



# Replace with the path to your Excel file
excel_file = 'C:\Python\Repository\Mobax\Device-Inventory.xlsx'

# Replace with the sheet name that contains the data
#sheet_name = 'Device-Inventory'
from openpyxl import Workbook, load_workbook
wb = load_workbook(excel_file)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    sheet_name = ws.title

    
# Replace with the names of the columns that contain the Device and IP data
device_col = 'Device Name'
ip_col = 'IP Address'

# Replace with the desired output file name
output_file = 'C:\Python\Repository\Mobax\output.txt'

# Load the Excel data into a pandas dataframe
df = pd.read_excel(excel_file, sheet_name=sheet_name)

# Extract the Device and IP columns and convert them to HTML entity codes
devices = df[device_col].apply(lambda x: x.replace(' ', '%20').replace('#', '%23'))
ips = df[ip_col].apply(lambda x: x.replace(' ', '%20').replace('#', '%23'))

# Generate the output strings
output = [format(device) + '=#109#0%' + format(ip) + '%22%[DNAC Admin]%%-1%-1%%%%%0%0%0%%%-1%0%0%0%%1080%%0%0%1#MobaFont%10%0%0%-1%15%4,173,7%30,30,30%180,180,192%0%-1%0%<>%xterm%-1%-1%_Std_Colors_0_%80%24%0%1%-1%<none>%%0%0%-1#0# #-1' + '\n'
         for device, ip in zip(devices, ips)]

# Write the output to a file
with open(output_file, 'w') as f:
    f.writelines(output)